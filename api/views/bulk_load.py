#!/usr/bin/python3
"""
RestFul API bulk_load for users
"""
from flask import jsonify, request
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
from api.views import app_views
from models.user import search_fullname_id
from models import storage
import os


UPLOAD_FOLDER = os.path.abspath("./uploads/")
ALLOWED_EXTENSIONS = set(["xlsx"])


def allowed_file(filname):
    """
    Method to validate the allowed extensions
    """
    if "." in filname and filname.split(".")[1].lower() in ALLOWED_EXTENSIONS:
        return True
    return False


@app_views.route("/bulkload", methods=["POST"])
def bulk_load():
    """
    Endpoint that loads the excel file to then carry out
    the bulk load of user and income data
    """
    if request.method == "POST":
        print("request.files {}".format(request.files))
        f = request.files["file"]
        if f.filename == "":
            return "No file selected."
        if f and allowed_file(f.filename):
            filename = secure_filename(f.filename).lower()
            f.save(os.path.join(UPLOAD_FOLDER, filename))
            get_file = os.scandir(UPLOAD_FOLDER)
            with get_file as files_folder:
                for fichero in files_folder:
                    if fichero.name == 'carga_masiva_user.xlsx':
                        path = UPLOAD_FOLDER+'/'+fichero.name
                        workbook = load_workbook(path)
                        sheet = workbook.active
                        listaUser = []
                        for r in range(2, sheet.max_row + 1):
                            values = (
                                    sheet.cell(r, 1).value,
                                    sheet.cell(r, 2).value,
                                    storage.get_id('document_type', 'name',
                                                   sheet.cell(r, 3).value),
                                    sheet.cell(r, 4).value,
                                    sheet.cell(r, 5).value,
                                    sheet.cell(r, 6).value,
                                    sheet.cell(r, 7).value,
                                    sheet.cell(r, 8).value,
                                    storage.get_id('role', 'name',
                                                   sheet.cell(r, 9).value)
                                )
                            listaUser.append(values)
                        storage.create_many(listaUser, 'usuarios')
                        os.remove(path)
                    elif fichero.name == 'carga_masiva_ingresos.xlsx':
                        path = UPLOAD_FOLDER+'/'+fichero.name
                        workbook = load_workbook(path)
                        sheet = workbook.active
                        listaIngresos = []
                        for row in range(2, sheet.max_row + 1):
                            values = (
                                    sheet.cell(row, 1).value,
                                    search_fullname_id(
                                        sheet.cell(row, 2).value),
                                    storage.get_id('departament',
                                                   'number_departament',
                                                   sheet.cell(row, 3).value),
                                    storage.get_id('type_payment', 'name_type',
                                                   sheet.cell(row, 4).value),
                                    sheet.cell(row, 5).value,
                                    sheet.cell(row, 6).value,
                                    sheet.cell(row, 7).value,
                                    sheet.cell(row, 8).value,
                                    sheet.cell(row, 9).value,
                                    sheet.cell(row, 10).value,
                                    sheet.cell(row, 11).value,
                                    storage.get_id('concept', 'name_concept',
                                                   sheet.cell(row, 12).value),
                                    sheet.cell(row, 13).value,
                                    sheet.cell(row, 14).value
                                )
                            listaIngresos.append(values)
                        storage.create_many(listaIngresos, 'ingresos')
                        os.remove(path)
        else:
            return jsonify({'mensaje': 'File not allowed'}), 406
    return jsonify({"status": "OK"}), 201
